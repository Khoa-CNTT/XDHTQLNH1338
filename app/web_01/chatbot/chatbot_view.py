import json
import os
import pandas as pd
from datetime import datetime, timedelta
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Count, F
from django.utils import timezone
from django.conf import settings
from django.urls import reverse
    # Tạo workbook Excel
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from web_01.models import (
    ChatHistory, Invoice, Order, OrderDetail, 
    Ingredient, InventoryLog, Table, Session, Product
)
from web_01.chatbot.gemini_service import GeminiChatbot

@login_required
def chatbot_view(request):
    """Hiển thị giao diện chatbot"""
    # Lấy lịch sử chat gần đây
    recent_chats = ChatHistory.objects.all().order_by('-created_at')[:10]
    
    context = {
        'recent_chats': recent_chats
    }
    
    return render(request, 'apps/web_01/dashboard/chatbot.html', context)

@csrf_exempt
@login_required
def chatbot_api(request):
    """API endpoint cho chatbot"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body) if request.body else {}
            if not data and request.POST:
                # Xử lý form data
                user_message = request.POST.get('message', '')
            else:
                user_message = data.get('message', '')
            
            if not user_message:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Không có tin nhắn'
                }, status=400)
            
            # Khởi tạo chatbot
            chatbot = GeminiChatbot()
            
            # Kiểm tra nếu yêu cầu xuất báo cáo Excel
            if "xuất báo cáo" in user_message.lower() or "excel" in user_message.lower():
                # Xác định loại báo cáo và khoảng thời gian
                report_type, period = extract_report_info(user_message)
                
                # Tạo và trả về file Excel
                if report_type and period:
                    excel_file, filename = generate_excel_report(report_type, period)
                    report_type_vi,period_vi= convert_en_to_vi_title(report_type, period)
                    # Lưu lịch sử chat với thông báo về báo cáo đã tạo
                    bot_reply = f"""
# Báo cáo Excel đã được tạo

Bot RMS 65 đã tạo báo cáo Excel **{report_type_vi}** cho **{get_period_name(period_vi)}**.

Bạn có thể tải xuống báo cáo bằng cách nhấp vào [đường dẫn này](/download-report/?type={report_type}&period={period}).

Báo cáo bao gồm:
- Tổng quan doanh thu
- Chi tiết theo sản phẩm
- Biểu đồ phân tích
                    """
                    
                    chatbot.save_chat_history(user_message, bot_reply)
                    
                    return JsonResponse({
                        'status': 'success',
                        'reply': bot_reply,
                        'report_url': f"/download-report/?type={report_type}&period={period}"
                    })
            
            # Xử lý câu hỏi thông thường
            bot_reply = chatbot.process_query(user_message)
            
            # Lưu lịch sử chat
            chatbot.save_chat_history(user_message, bot_reply)
            
            return JsonResponse({
                'status': 'success',
                'reply': bot_reply
            })
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'status': 'error',
        'message': 'Method not allowed'
    }, status=405)

@login_required
def get_chat_history(request):
    """API endpoint để lấy lịch sử chat"""
    try:
        # Lấy 10 cuộc trò chuyện gần nhất
        chat_history = ChatHistory.objects.all().order_by('-created_at')[:10]
        
        # Đảo ngược để hiển thị theo thứ tự thời gian
        chat_history = reversed(list(chat_history))
        
        # Chuyển đổi thành JSON
        history_data = []
        for chat in chat_history:
            history_data.append({
                'user_message': chat.user_message,
                'bot_reply': chat.bot_reply,
                'created_at': chat.created_at.strftime('%d/%m/%Y %H:%M:%S')
            })
        
        return JsonResponse({
            'status': 'success',
            'chat_history': history_data
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def download_report(request):
    """Tải xuống báo cáo Excel"""
    report_type = request.GET.get('type', 'revenue')
    period = request.GET.get('period', 'today')
    
    excel_file, filename = generate_excel_report(report_type, period)
    
    response = HttpResponse(
        excel_file,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def extract_report_info(message):
    """Trích xuất loại báo cáo và khoảng thời gian từ tin nhắn"""
    message = message.lower()
    
    # Xác định loại báo cáo
    report_type = 'revenue'  # Mặc định là báo cáo doanh thu
    if "tồn kho" in message or "nguyên liệu" in message:
        report_type = 'inventory'
    elif "sản phẩm" in message or "món ăn" in message:
        report_type = 'products'
    elif "bàn" in message:
        report_type = 'tables'
    
    # Xác định khoảng thời gian
    period = 'today'  # Mặc định là hôm nay
    if "hôm qua" in message:
        period = 'yesterday'
    elif "tuần" in message:
        period = 'week'
    elif "tháng" in message:
        period = 'month'
    elif "năm" in message:
        period = 'year'
    
    return report_type, period


def convert_en_to_vi_title(report_type, period):
    """Chuyển đổi report_type và period từ tiếng Anh sang tiếng Việt dạng title case"""
    # Từ điển ánh xạ report_type sang tiếng Việt
    report_type_map = {
        'revenue': 'Doanh Thu',
        'inventory': 'Tồn Kho',
        'products': 'Sản Phẩm',
        'tables': 'Bàn'
    }
    
    # Từ điển ánh xạ period sang tiếng Việt
    period_map = {
        'today': 'Hôm Nay',
        'yesterday': 'Hôm Qua',
        'week': 'Tuần',
        'month': 'Tháng',
        'year': 'Năm'
    }
    
    # Lấy giá trị tiếng Việt, mặc định giữ nguyên nếu không tìm thấy
    report_type_vi = report_type_map.get(report_type, report_type.title())
    period_vi = period_map.get(period, period.title())
    
    return report_type_vi, period_vi

def get_period_name(period):
    """Trả về tên khoảng thời gian dễ đọc"""
    period_names = {
        'today': 'hôm nay',
        'yesterday': 'hôm qua',
        'week': '7 ngày qua',
        'month': 'tháng này',
        'year': 'năm nay'
    }
    return period_names.get(period, period)

def generate_excel_report(report_type, period):
    """Tạo báo cáo Excel dựa trên loại và khoảng thời gian"""
    today = timezone.now().date()
    
    # Xác định khoảng thời gian
    if period == 'today':
        start_date = today
        end_date = today
    elif period == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    else:
        start_date = today
        end_date = today
    

    
    wb = Workbook()
    ws = wb.active
    
    # Định dạng tiêu đề
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=12, bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Tạo báo cáo dựa trên loại
    if report_type == 'revenue':
        ws.title = "Báo Cáo Doanh Thu"
        
        # Tiêu đề báo cáo
        ws['A1'] = f"BÁO CÁO DOANH THU - {get_period_name(period).upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Thêm ngày tạo báo cáo
        ws['A2'] = f"Ngày tạo: {today.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='left')
        ws.merge_cells('A2:G2')
        
        # Tiêu đề cột
        headers = ["STT", "Ngày", "Số đơn hàng", "Doanh thu", "Giá trị TB/đơn", "Sản phẩm bán chạy", "Số lượng"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu doanh thu
        row = 5
        stt = 1
        
        # Nếu là báo cáo theo ngày
        if period in ['today', 'yesterday']:
            # Doanh thu theo giờ
            for hour in range(24):
                hour_start = datetime.combine(start_date, datetime.min.time()) + timedelta(hours=hour)
                hour_end = hour_start + timedelta(hours=1)
                
                # Đảm bảo timezone nhất quán
                if settings.USE_TZ:
                    from django.utils.timezone import make_aware
                    hour_start = make_aware(hour_start)
                    hour_end = make_aware(hour_end)
                
                # Doanh thu
                revenue = Invoice.objects.filter(
                    created_at__gte=hour_start,
                    created_at__lt=hour_end,
                    is_deleted=False
                ).aggregate(total=Sum('total_amount'))['total'] or 0
                
                # Số đơn hàng
                orders = Order.objects.filter(
                    created_at__gte=hour_start,
                    created_at__lt=hour_end,
                    is_deleted=False
                )
                order_count = orders.count()
                
                # Nếu không có đơn hàng trong giờ này, bỏ qua
                if order_count == 0:
                    continue
                
                # Doanh thu trung bình mỗi đơn
                avg_order_value = revenue / order_count if order_count > 0 else 0
                
                # Món bán chạy nhất trong giờ này
                best_selling = OrderDetail.objects.filter(
                    created_at__gte=hour_start,
                    created_at__lt=hour_end,
                    is_deleted=False
                ).values('product__name').annotate(
                    total_sold=Sum('quantity')
                ).order_by('-total_sold').first()
                
                best_product = best_selling['product__name'] if best_selling else "Không có"
                best_quantity = best_selling['total_sold'] if best_selling else 0
                
                # Thêm dữ liệu vào báo cáo
                ws.cell(row=row, column=1, value=stt)
                ws.cell(row=row, column=2, value=f"{hour_start.strftime('%H:%M')} - {hour_end.strftime('%H:%M')}")
                ws.cell(row=row, column=3, value=order_count)
                ws.cell(row=row, column=4, value=revenue)
                ws.cell(row=row, column=5, value=avg_order_value)
                ws.cell(row=row, column=6, value=best_product)
                ws.cell(row=row, column=7, value=best_quantity)
                
                row += 1
                stt += 1
        else:
            # Báo cáo theo ngày trong khoảng thời gian
            current_date = start_date
            while current_date <= end_date:
                # Doanh thu
                revenue = Invoice.objects.filter(
                    created_at__date=current_date,
                    is_deleted=False
                ).aggregate(total=Sum('total_amount'))['total'] or 0
                
                # Số đơn hàng
                orders = Order.objects.filter(
                    created_at__date=current_date,
                    is_deleted=False
                )
                order_count = orders.count()
                
                # Doanh thu trung bình mỗi đơn
                avg_order_value = revenue / order_count if order_count > 0 else 0
                
                # Món bán chạy nhất trong ngày
                best_selling = OrderDetail.objects.filter(
                    created_at__date=current_date,
                    is_deleted=False
                ).values('product__name').annotate(
                    total_sold=Sum('quantity')
                ).order_by('-total_sold').first()
                
                best_product = best_selling['product__name'] if best_selling else "Không có"
                best_quantity = best_selling['total_sold'] if best_selling else 0
                
                # Thêm dữ liệu vào báo cáo
                ws.cell(row=row, column=1, value=stt)
                ws.cell(row=row, column=2, value=current_date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=3, value=order_count)
                ws.cell(row=row, column=4, value=revenue)
                ws.cell(row=row, column=5, value=avg_order_value)
                ws.cell(row=row, column=6, value=best_product)
                ws.cell(row=row, column=7, value=best_quantity)
                
                row += 1
                stt += 1
                current_date += timedelta(days=1)
        
        # Tổng cộng
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="TỔNG CỘNG")
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        # Tính tổng số đơn hàng
        total_orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            is_deleted=False
        ).count()
        ws.cell(row=row, column=3, value=total_orders)
        
        # Tính tổng doanh thu
        total_revenue = Invoice.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            is_deleted=False
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        ws.cell(row=row, column=4, value=total_revenue)
        
        # Tính giá trị trung bình mỗi đơn
        avg_value = total_revenue / total_orders if total_orders > 0 else 0
        ws.cell(row=row, column=5, value=avg_value)
        
        # Định dạng cột
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Thêm sheet phân tích sản phẩm
        ws_products = wb.create_sheet(title="Phân Tích Sản Phẩm")
        
        # Tiêu đề
        ws_products['A1'] = f"PHÂN TÍCH SẢN PHẨM - {get_period_name(period).upper()}"
        ws_products['A1'].font = title_font
        ws_products.merge_cells('A1:E1')
        ws_products['A1'].alignment = Alignment(horizontal='center')
        
        # Tiêu đề cột
        product_headers = ["STT", "Sản phẩm", "Số lượng bán", "Doanh thu", "Tỷ lệ"]
        for col, header in enumerate(product_headers, 1):
            cell = ws_products.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu sản phẩm bán chạy
        best_selling_products = OrderDetail.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            is_deleted=False
        ).values('product__name').annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(F('price') * F('quantity'))
        ).order_by('-total_sold')[:20]
        
        # Tính tổng doanh thu từ tất cả sản phẩm
        total_product_revenue = sum(product['revenue'] for product in best_selling_products)
        
        # Thêm dữ liệu vào báo cáo
        row = 4
        for idx, product in enumerate(best_selling_products, 1):
            percentage = (product['revenue'] / total_product_revenue * 100) if total_product_revenue > 0 else 0
            
            ws_products.cell(row=row, column=1, value=idx)
            ws_products.cell(row=row, column=2, value=product['product__name'])
            ws_products.cell(row=row, column=3, value=product['total_sold'])
            ws_products.cell(row=row, column=4, value=product['revenue'])
            ws_products.cell(row=row, column=5, value=f"{percentage:.2f}%")
            
            row += 1
        
        # Định dạng cột
        for col in range(1, 6):
            ws_products.column_dimensions[get_column_letter(col)].width = 15
        
    elif report_type == 'inventory':
        ws.title = "Báo Cáo Tồn Kho"
        
        # Tiêu đề báo cáo
        ws['A1'] = f"BÁO CÁO TỒN KHO NGUYÊN LIỆU - {today.strftime('%d/%m/%Y')}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Tiêu đề cột
        headers = ["STT", "Nguyên liệu", "Đơn vị", "Số lượng tồn", "Nhập gần nhất", "Xuất gần nhất"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu tồn kho
        ingredients = Ingredient.objects.all().order_by('name')
        
        row = 4
        for idx, ingredient in enumerate(ingredients, 1):
            # Tìm lần nhập gần nhất
            last_import = InventoryLog.objects.filter(
                ingredient=ingredient,
                change__gt=0
            ).order_by('-last_updated').first()
            
            # Tìm lần xuất gần nhất
            last_export = InventoryLog.objects.filter(
                ingredient=ingredient,
                change__lt=0
            ).order_by('-last_updated').first()
            
            # Thêm dữ liệu vào báo cáo
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=ingredient.name)
            ws.cell(row=row, column=3, value=ingredient.unit)
            ws.cell(row=row, column=4, value=ingredient.quantity_in_stock)
            
            if last_import:
                ws.cell(row=row, column=5, value=f"{last_import.last_updated.strftime('%d/%m/%Y %H:%M')} ({last_import.change} {ingredient.unit})")
            else:
                ws.cell(row=row, column=5, value="Chưa có")
            
            if last_export:
                ws.cell(row=row, column=6, value=f"{last_export.last_updated.strftime('%d/%m/%Y %H:%M')} ({abs(last_export.change)} {ingredient.unit})")
            else:
                ws.cell(row=row, column=6, value="Chưa có")
            
            row += 1
        
        # Định dạng cột
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Thêm sheet nguyên liệu sắp hết
        ws_low = wb.create_sheet(title="Nguyên Liệu Sắp Hết")
        
        # Tiêu đề
        ws_low['A1'] = "DANH SÁCH NGUYÊN LIỆU SẮP HẾT"
        ws_low['A1'].font = title_font
        ws_low.merge_cells('A1:D1')
        ws_low['A1'].alignment = Alignment(horizontal='center')
        
        # Tiêu đề cột
        low_headers = ["STT", "Nguyên liệu", "Đơn vị", "Số lượng tồn"]
        for col, header in enumerate(low_headers, 1):
            cell = ws_low.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy danh sách nguyên liệu sắp hết
        low_stock = []
        for ingredient in ingredients:
            if hasattr(ingredient, 'min_stock') and ingredient.quantity_in_stock <= ingredient.min_stock:
                low_stock.append(ingredient)
            elif ingredient.quantity_in_stock <= 50:  # Ngưỡng mặc định
                low_stock.append(ingredient)
        
        # Thêm dữ liệu vào báo cáo
        row = 4
        for idx, ingredient in enumerate(low_stock, 1):
            ws_low.cell(row=row, column=1, value=idx)
            ws_low.cell(row=row, column=2, value=ingredient.name)
            ws_low.cell(row=row, column=3, value=ingredient.unit)
            ws_low.cell(row=row, column=4, value=ingredient.quantity_in_stock)
            
            row += 1
        
        # Định dạng cột
        for col in range(1, 5):
            ws_low.column_dimensions[get_column_letter(col)].width = 15
    
    elif report_type == 'products':
        ws.title = "Báo Cáo Sản Phẩm"
        
        # Tiêu đề báo cáo
        ws['A1'] = f"BÁO CÁO SẢN PHẨM - {get_period_name(period).upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Tiêu đề cột
        headers = ["STT", "Sản phẩm", "Danh mục", "Giá", "Số lượng bán", "Doanh thu", "Tỷ lệ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu sản ph��m
        products = Product.objects.filter(is_deleted=False).order_by('category__name', 'name')
        
        # Lấy dữ liệu bán hàng trong khoảng thời gian
        sales_data = OrderDetail.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            is_deleted=False
        ).values('product').annotate(
            total_sold=Sum('quantity'),
            revenue=Sum(F('price') * F('quantity'))
        )
        
        # Tạo dictionary để tra cứu nhanh
        sales_dict = {item['product']: item for item in sales_data}
        
        # Tính tổng doanh thu
        total_revenue = sum(item['revenue'] for item in sales_data)
        
        # Thêm dữ liệu vào báo cáo
        row = 4
        for idx, product in enumerate(products, 1):
            sales_info = sales_dict.get(product.id, {'total_sold': 0, 'revenue': 0})
            percentage = (sales_info['revenue'] / total_revenue * 100) if total_revenue > 0 and sales_info['revenue'] > 0 else 0
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=product.name)
            ws.cell(row=row, column=3, value=product.category.name if product.category else "")
            ws.cell(row=row, column=4, value=product.price)
            ws.cell(row=row, column=5, value=sales_info['total_sold'])
            ws.cell(row=row, column=6, value=sales_info['revenue'])
            ws.cell(row=row, column=7, value=f"{percentage:.2f}%")
            
            row += 1
        
        # Định dạng cột
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Thêm sheet phân tích danh mục
        ws_cat = wb.create_sheet(title="Phân Tích Danh Mục")
        
        # Tiêu đề
        ws_cat['A1'] = f"PHÂN TÍCH DANH MỤC - {get_period_name(period).upper()}"
        ws_cat['A1'].font = title_font
        ws_cat.merge_cells('A1:E1')
        ws_cat['A1'].alignment = Alignment(horizontal='center')
        
        # Tiêu đề cột
        cat_headers = ["STT", "Danh mục", "Số sản phẩm", "Doanh thu", "Tỷ lệ"]
        for col, header in enumerate(cat_headers, 1):
            cell = ws_cat.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu theo danh mục
        category_data = OrderDetail.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            is_deleted=False
        ).values('product__category__name').annotate(
            product_count=Count('product', distinct=True),
            revenue=Sum(F('price') * F('quantity'))
        ).order_by('-revenue')
        
        # Thêm dữ liệu vào báo cáo
        row = 4
        for idx, category in enumerate(category_data, 1):
            cat_name = category['product__category__name'] or "Không có danh mục"
            percentage = (category['revenue'] / total_revenue * 100) if total_revenue > 0 else 0
            
            ws_cat.cell(row=row, column=1, value=idx)
            ws_cat.cell(row=row, column=2, value=cat_name)
            ws_cat.cell(row=row, column=3, value=category['product_count'])
            ws_cat.cell(row=row, column=4, value=category['revenue'])
            ws_cat.cell(row=row, column=5, value=f"{percentage:.2f}%")
            
            row += 1
        
        # Định dạng cột
        for col in range(1, 6):
            ws_cat.column_dimensions[get_column_letter(col)].width = 15
    
    elif report_type == 'tables':
        ws.title = "Báo Cáo Bàn"
        
        # Tiêu đề báo cáo
        ws['A1'] = f"BÁO CÁO SỬ DỤNG BÀN - {get_period_name(period).upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Tiêu đề cột
        headers = ["STT", "Bàn số", "Số phiên", "Thời gian TB (phút)", "Doanh thu", "Doanh thu TB/phiên"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Lấy dữ liệu bàn
        tables = Table.objects.filter(is_deleted=False).order_by('table_number')
        
        row = 4
        for idx, table in enumerate(tables, 1):
            # Lấy các phiên trong khoảng thời gian
            sessions = Session.objects.filter(
                table=table,
                started_at__date__gte=start_date,
                started_at__date__lte=end_date
            )
            
            session_count = sessions.count()
            
            # Tính thời gian trung bình của phiên
            avg_duration = 0
            completed_sessions = sessions.filter(status='closed', ended_at__isnull=False)
            
            if completed_sessions.exists():
                total_duration = 0
                for session in completed_sessions:
                    duration = (session.ended_at - session.started_at).total_seconds() / 60  # Phút
                    total_duration += duration
                
                avg_duration = total_duration / completed_sessions.count()
            
            # Tính doanh thu từ các phiên
            session_ids = sessions.values_list('id', flat=True)
            revenue = Invoice.objects.filter(
                session_id__in=session_ids,
                is_deleted=False
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            
            # Doanh thu trung bình mỗi phiên
            avg_revenue = revenue / session_count if session_count > 0 else 0
            
            # Thêm dữ liệu vào báo cáo
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=f"Bàn {table.table_number}")
            ws.cell(row=row, column=3, value=session_count)
            ws.cell(row=row, column=4, value=f"{avg_duration:.1f}")
            ws.cell(row=row, column=5, value=revenue)
            ws.cell(row=row, column=6, value=avg_revenue)
            
            row += 1
        
        # Định dạng cột
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Lưu workbook vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Tạo tên file
    filename = f"{report_type}_{period}_{today.strftime('%Y%m%d')}.xlsx"
    
    return buffer, filename
